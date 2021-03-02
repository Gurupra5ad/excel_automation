from openpyxl import load_workbook

book = load_workbook('main.xlsx')
sheet = book.active

new_data = [1500, 12500, 675, 1340, 12450]

for row in sheet['b5':'f5']:
    for index, cell in enumerate(row):
        cell.value = new_data[index]
    
book.save('new_report.xlsx')

new_read = load_workbook('new_report.xlsx')
sheet = book.active

for row in sheet:
    for cell in row:
        print(cell.value)